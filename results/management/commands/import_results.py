import openpyxl

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from users.models import CustomUser
from stages.models import Family

from results.models import (
    Subject,
    Exam,
    SubjectExam,
    StudentEnrollment,
    Result,
)


class Command(BaseCommand):

    help = "Import annual results from Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Path to Excel file"
        )

        parser.add_argument(
            "--family",
            required=True,
            help="Family name"
        )

        parser.add_argument(
            "--academic-year",
            required=True,
            help="Academic year, example: 2025/2026"
        )

    @transaction.atomic
    def handle(self, *args, **options):

        file_path = options["file"]
        family_name = options["family"]
        academic_year = options["academic_year"]

        # ==========================================
        # Load Excel
        # ==========================================

        try:
            workbook = openpyxl.load_workbook(
                file_path,
                data_only=True
            )
        except Exception as e:
            raise CommandError(
                f"Unable to open Excel file: {e}"
            )

        if "Sheet1" not in workbook.sheetnames:
            raise CommandError(
                "Sheet1 was not found in the Excel file."
            )

        sheet = workbook["Sheet1"]

        # ==========================================
        # Family
        # ==========================================

        try:
            family = Family.objects.get(
                name=family_name
            )
        except Family.DoesNotExist:
            raise CommandError(
                f'Family "{family_name}" does not exist.'
            )

        # ==========================================
        # Subjects
        # ==========================================

        subjects_config = {

            "الألحان": {
                "final_grade": 300,
                "success_grade": 150,
            },

            "طقس": {
                "final_grade": 40,
                "success_grade": 20,
            },

            "قبطي": {
                "final_grade": 40,
                "success_grade": 20,
            },

            "كتاب مقدس": {
                "final_grade": 30,
                "success_grade": 15,
            },

            "عقيدة": {
                "final_grade": 25,
                "success_grade": 12.5,
            },

            "أجبية": {
                "final_grade": 15,
                "success_grade": 7.5,
            },

            "الحضور والغياب": {
                "final_grade": 150,
                "success_grade": 75,
            },
        }

        subjects = {}

        for name, config in subjects_config.items():

            subject, created = Subject.objects.get_or_create(
                name=name,
                defaults=config
            )

            if not created:

                changed = False

                if subject.final_grade != config["final_grade"]:
                    subject.final_grade = config["final_grade"]
                    changed = True

                if subject.success_grade != config["success_grade"]:
                    subject.success_grade = config["success_grade"]
                    changed = True

                if changed:
                    subject.save()

            subjects[name] = subject

        # ==========================================
        # Exams
        # ==========================================

        exams_config = {
            "ترم 1": 100,
            "ترم 2": 110,
            "ترم 3": 90,
        }

        exams = {}

        for name in exams_config:

            exam, _ = Exam.objects.get_or_create(
                name=name,
                year=academic_year
            )

            exams[name] = exam

        # ==========================================
        # Subject Exams
        # ==========================================

        # max_grade = درجة المادة في هذا الترم
        #
        # success_grade = نصف درجة الترم
        #
        # وليس success_grade السنوي الموجود في Subject

        subject_exam_config = {

            # ======================================
            # الألحان
            # ======================================

            ("الألحان", "ترم 1"): {
                "max_grade": 100,
                "success_grade": 50,
            },

            ("الألحان", "ترم 2"): {
                "max_grade": 110,
                "success_grade": 55,
            },

            ("الألحان", "ترم 3"): {
                "max_grade": 90,
                "success_grade": 45,
            },

            # ======================================
            # مواد ترم واحد
            # ======================================

            ("طقس", "ترم 1"): {
                "max_grade": 40,
                "success_grade": 20,
            },

            ("قبطي", "ترم 2"): {
                "max_grade": 40,
                "success_grade": 20,
            },

            ("كتاب مقدس", "ترم 3"): {
                "max_grade": 30,
                "success_grade": 15,
            },

            ("عقيدة", "ترم 3"): {
                "max_grade": 25,
                "success_grade": 12.5,
            },

            # ======================================
            # الأجبية
            # ======================================

            ("أجبية", "ترم 1"): {
                "max_grade": 15,
                "success_grade": 7.5,
            },

            # ======================================
            # الحضور والغياب
            # ======================================

            ("الحضور والغياب", "ترم 1"): {
                "max_grade": 50,
                "success_grade": 25,
            },

            ("الحضور والغياب", "ترم 2"): {
                "max_grade": 50,
                "success_grade": 25,
            },

            ("الحضور والغياب", "ترم 3"): {
                "max_grade": 50,
                "success_grade": 25,
            },
        }

        subject_exams = {}

        for (
            subject_name,
            exam_name
        ), config in subject_exam_config.items():

            subject_exam, _ = SubjectExam.objects.update_or_create(
                subject=subjects[subject_name],
                exam=exams[exam_name],
                defaults={
                    "max_grade": config["max_grade"],
                    "success_grade": config["success_grade"],
                }
            )

            subject_exams[
                (subject_name, exam_name)
            ] = subject_exam

        # ==========================================
        # Subject Components
        # ==========================================

        components = {}

        for component_name in [
            "المزمور الأول",
            "المزمور الثاني",
            "المزمور الثالث",
        ]:

            component, _ = SubjectComponent.objects.get_or_create(
                subject=subjects["أجبية"],
                name=component_name
            )

            components[component_name] = component

        # ==========================================
        # Component Exams
        # ==========================================

        component_exams = {}

        for component_name, component in components.items():

            component_exam, _ = ComponentExam.objects.update_or_create(
                component=component,
                exam=exams["ترم 1"],
                defaults={
                    "max_grade": 5,
                    "success_grade": 2.5,
                }
            )

            component_exams[
                component_name
            ] = component_exam

        # ==========================================
        # Students
        # ==========================================

        imported_students = 0
        imported_results = 0

        for row in range(5, sheet.max_row + 1):

            student_number = sheet.cell(row, 1).value
            student_name = sheet.cell(row, 2).value

            # ======================================
            # Skip empty rows
            # ======================================

            if not student_name:
                continue

            student_name = str(student_name).strip()

            # ======================================
            # Find student
            # ======================================

            try:
                student = CustomUser.objects.get(
                    username=student_name
                )

            except CustomUser.DoesNotExist:

                raise CommandError(
                    f'Student "{student_name}" was not found.'
                )

            # ======================================
            # Enrollment
            # ======================================

            enrollment, _ = StudentEnrollment.objects.get_or_create(
                student=student,
                family=family,
                academic_year=academic_year
            )

            imported_students += 1

            # ======================================
            # Helper
            # ======================================

            def create_subject_result(
                subject_name,
                exam_name,
                points
            ):

                if points is None:
                    return

                subject_exam = subject_exams[
                    (subject_name, exam_name)
                ]

                Result.objects.update_or_create(
                    enrollment=enrollment,
                    subject_exam=subject_exam,
                    defaults={
                        "component_exam": None,
                        "points": float(points)
                    }
                )

            def create_component_result(
                component_name,
                points
            ):

                if points is None:
                    return

                component_exam = component_exams[
                    component_name
                ]

                Result.objects.update_or_create(
                    enrollment=enrollment,
                    component_exam=component_exam,
                    defaults={
                        "subject_exam": None,
                        "points": float(points)
                    }
                )

            # ======================================
            # الألحان
            # ======================================

            create_subject_result(
                "الألحان",
                "ترم 1",
                sheet.cell(row, 3).value
            )

            create_subject_result(
                "الألحان",
                "ترم 2",
                sheet.cell(row, 4).value
            )

            create_subject_result(
                "الألحان",
                "ترم 3",
                sheet.cell(row, 5).value
            )

            # ======================================
            # المواد الدراسية
            # ======================================

            create_subject_result(
                "طقس",
                "ترم 1",
                sheet.cell(row, 7).value
            )

            create_subject_result(
                "قبطي",
                "ترم 2",
                sheet.cell(row, 8).value
            )

            create_subject_result(
                "كتاب مقدس",
                "ترم 3",
                sheet.cell(row, 9).value
            )

            create_subject_result(
                "عقيدة",
                "ترم 3",
                sheet.cell(row, 10).value
            )

            # ======================================
            # الأجبية
            # ======================================

            create_component_result(
                "المزمور الأول",
                sheet.cell(row, 11).value
            )

            create_component_result(
                "المزمور الثاني",
                sheet.cell(row, 12).value
            )

            create_component_result(
                "المزمور الثالث",
                sheet.cell(row, 13).value
            )

            # ======================================
            # الحضور والغياب
            # ======================================

            create_subject_result(
                "الحضور والغياب",
                "ترم 1",
                sheet.cell(row, 15).value
            )

            create_subject_result(
                "الحضور والغياب",
                "ترم 2",
                sheet.cell(row, 16).value
            )

            create_subject_result(
                "الحضور والغياب",
                "ترم 3",
                sheet.cell(row, 17).value
            )

            # ======================================
            # Calculate annual status
            # ======================================

            from results.services import update_annual_status

            update_annual_status(enrollment)

            imported_results += 1

        # ==========================================
        # Done
        # ==========================================

        self.stdout.write(
            self.style.SUCCESS(
                "Import completed successfully."
            )
        )

        self.stdout.write(
            f"Students imported: {imported_students}"
        )

        self.stdout.write(
            f"Enrollments processed: {imported_results}"
        )