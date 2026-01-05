import re
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth.hashers import make_password
from users.models import CustomUser
from stages.models import Stage, Family, Child, Servant
import openpyxl

DEFAULT_PW_SUFFIX = "@123"
CHILD_SERIAL_START = 1001
SERVANT_SERIAL_START = 2001
DEFAULT_PASSWORD = "TempPass123"  # will be overridden by pattern username@123

def safe_str(x):
    return str(x).strip() if x is not None else ""

def split_first_last(full_name):
    parts = safe_str(full_name).split()
    first_name = parts[0].capitalize() if len(parts) > 0 else "User"
    last_name = parts[1].capitalize() if len(parts) > 1 else ""
    return first_name, last_name

def next_serial_for(first_name_prefix, kind):
    """
    kind = 'child' or 'servant'
    We'll scan usernames starting with prefix and extract the numeric suffix, then return next.
    """
    prefix = f"{first_name_prefix}{'C' if kind=='child' else 'S'}"
    qs = CustomUser.objects.filter(username__startswith=prefix)
    max_num = 0
    regex = re.compile(rf'^{re.escape(prefix)}(\d+)$')
    for u in qs:
        m = regex.match(u.username)
        if m:
            try:
                num = int(m.group(1))
                if num > max_num:
                    max_num = num
            except ValueError:
                continue
    if max_num:
        return max_num + 1
    else:
        return CHILD_SERIAL_START if kind == 'child' else SERVANT_SERIAL_START

def generate_username(first_name, kind):
    serial = next_serial_for(first_name, kind)
    prefix = 'C' if kind == 'child' else 'S'
    return f"{first_name}{prefix}{serial}"

def generate_password(username):
    return f"{username}{DEFAULT_PW_SUFFIX}"

class Command(BaseCommand):
    help = "Seed Stage, Family, Child, Servant from Excel files: stage.xlsx, family.xlsx, children.xlsx, servants.xlsx"

    def handle(self, *args, **options):
        summary = {"stages":0, "families":0, "children":0, "servants":0, "errors":[]}

        try:
            with transaction.atomic():
                # 1) stages
                try:
                    wb = openpyxl.load_workbook("stage.xlsx", data_only=True)
                    sheet = wb.active
                except FileNotFoundError:
                    self.stdout.write(self.style.ERROR("stage.xlsx not found."))
                    return

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    try:
                        stage_id = int(row[0]) if row[0] is not None else None
                    except Exception:
                        stage_id = None
                    name = safe_str(row[1])
                    if not name:
                        continue
                    if stage_id:
                        Stage.objects.update_or_create(id=stage_id, defaults={"name": name})
                    else:
                        Stage.objects.update_or_create(name=name, defaults={"name": name})
                    summary["stages"] += 1

                # 2) families
                try:
                    wb = openpyxl.load_workbook("family.xlsx", data_only=True)
                    sheet = wb.active
                except FileNotFoundError:
                    self.stdout.write(self.style.ERROR("family.xlsx not found."))
                    return

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    fam_id = row[0]
                    name = safe_str(row[1])
                    year = safe_str(row[2])
                    stage_id = row[3]
                    if not name:
                        summary["errors"].append(f"Family row missing name: {row}")
                        continue
                    try:
                        stage = Stage.objects.get(id=int(stage_id)) if stage_id else None
                    except Exception:
                        summary["errors"].append(f"Family '{name}' references missing stage id: {stage_id}")
                        continue
                    if fam_id:
                        Family.objects.update_or_create(
                            id=int(fam_id),
                            defaults={"name": name, "year": year, "stage": stage}
                        )
                    else:
                        Family.objects.update_or_create(
                            name=name,
                            defaults={"year": year, "stage": stage}
                        )
                    summary["families"] += 1

                # 3) children
                try:
                    wb = openpyxl.load_workbook("children.xlsx", data_only=True)
                    sheet = wb.active
                except FileNotFoundError:
                    self.stdout.write(self.style.ERROR("children.xlsx not found."))
                    return

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    # name | birth_date | phone | parent_phone | address | father | year_of_entrance | year_of_graduation | family_id
                    name = safe_str(row[0])
                    birth = row[1]
                    phone = safe_str(row[2])
                    parent_phone = safe_str(row[3])
                    address = safe_str(row[4])
                    father = safe_str(row[5])
                    year_in = row[6]
                    year_out = row[7]
                    family_id = row[8]

                    if not name:
                        summary["errors"].append(f"Child row missing name: {row}")
                        continue

                    try:
                        family = Family.objects.get(id=int(family_id)) if family_id else None
                    except Exception:
                        summary["errors"].append(f"Child '{name}' references missing family id: {family_id}")
                        continue

                    first_name, last_name = split_first_last(name)
                    # generate unique username
                    username = generate_username(first_name, kind='child')
                    password = generate_password(username)

                    # create or get user (if username collision extremely rare, loop to find next)
                    attempt = 0
                    while True:
                        try:
                            user = CustomUser.objects.create_user(
                                username=username,
                                password=password,
                                first_name=first_name,
                                last_name=last_name,
                                required_password_change=True
                            )
                            break
                        except Exception as e:
                            attempt += 1
                            username = generate_username(first_name, kind='child')
                            password = generate_password(username)
                            if attempt > 10:
                                summary["errors"].append(f"Failed creating user for child {name}: {e}")
                                user = None
                                break

                    if not user:
                        continue

                    # create child
                    try:
                        Child.objects.update_or_create(
                            user=user,
                            defaults={
                                "name": name,
                                "birth_date": birth,
                                "phone": phone,
                                "parent_phone": parent_phone,
                                "address": address,
                                "father": father,
                                "year_of_entrance": year_in,
                                "year_of_graduation": year_out,
                                "family": family
                            }
                        )
                        summary["children"] += 1
                    except Exception as e:
                        summary["errors"].append(f"Error creating Child for {name}: {e}")
                        # If child creation fails, delete the user to avoid orphan
                        try:
                            user.delete()
                        except:
                            pass
                        continue

                # 4) servants
                try:
                    wb = openpyxl.load_workbook("servants.xlsx", data_only=True)
                    sheet = wb.active
                except FileNotFoundError:
                    self.stdout.write(self.style.ERROR("servants.xlsx not found."))
                    return

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    # name | birth_date | address | role | stage_id | family_id
                    name = safe_str(row[0])
                    birth = row[1]
                    address = safe_str(row[2])
                    role = safe_str(row[3])
                    stage_id = row[4]
                    family_id = row[5]

                    if not name:
                        summary["errors"].append(f"Servant row missing name: {row}")
                        continue

                    try:
                        stage = Stage.objects.get(id=int(stage_id)) if stage_id else None
                    except Exception:
                        summary["errors"].append(f"Servant '{name}' references missing stage id: {stage_id}")
                        continue

                    try:
                        family = Family.objects.get(id=int(family_id)) if family_id else None
                    except Exception:
                        summary["errors"].append(f"Servant '{name}' references missing family id: {family_id}")
                        continue

                    first_name, last_name = split_first_last(name)
                    username = generate_username(first_name, kind='servant')
                    password = generate_password(username)

                    attempt = 0
                    while True:
                        try:
                            user = CustomUser.objects.create_user(
                                username=username,
                                password=password,
                                first_name=first_name,
                                last_name=last_name,
                                required_password_change=True
                            )
                            break
                        except Exception as e:
                            attempt += 1
                            username = generate_username(first_name, kind='servant')
                            password = generate_password(username)
                            if attempt > 10:
                                summary["errors"].append(f"Failed creating user for servant {name}: {e}")
                                user = None
                                break

                    if not user:
                        continue

                    try:
                        Servant.objects.update_or_create(
                            user=user,
                            defaults={
                                "name": name,
                                "birth_date": birth,
                                "address": address,
                                "role": role or "خادم عادي",
                                "stage": stage,
                                "family": family
                            }
                        )
                        summary["servants"] += 1
                    except Exception as e:
                        summary["errors"].append(f"Error creating Servant for {name}: {e}")
                        try:
                            user.delete()
                        except:
                            pass
                        continue

        except Exception as e_all:
            self.stdout.write(self.style.ERROR(f"Fatal error: {e_all}"))
            return

        # Final summary
        self.stdout.write(self.style.SUCCESS(f"Stages imported: {summary['stages']}"))
        self.stdout.write(self.style.SUCCESS(f"Families imported: {summary['families']}"))
        self.stdout.write(self.style.SUCCESS(f"Children imported: {summary['children']}"))
        self.stdout.write(self.style.SUCCESS(f"Servants imported: {summary['servants']}"))
        if summary["errors"]:
            self.stdout.write(self.style.WARNING("Some issues occurred:"))
            for err in summary["errors"]:
                self.stdout.write(self.style.WARNING(f" - {err}"))
